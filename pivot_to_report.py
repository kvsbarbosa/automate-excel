from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

month = 'february'

wb = load_workbook('pivot_table.xlsx')
sheet = wb['Report']

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# barchart initialization
bar_chart = BarChart()

# data and categories
data = Reference(sheet, min_col=min_column + 1, max_col=max_column, min_row=min_row, max_row=max_row)
categories = Reference(sheet, min_col=min_column, max_col=min_column, min_row=min_row + 1, max_row=max_row)

bar_chart.add_data(data, titles_from_data=True)
bar_chart.set_categories(categories)

sheet.add_chart(bar_chart, "B12")

bar_chart.title = "Sales by Product line"
bar_chart.style = 5

wb = load_workbook('pivot_table.xlsx')
sheet = wb['Report']

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

bar_chart = BarChart()

data = Reference(sheet, min_col=min_column + 1, max_col=max_column, min_row=min_row, max_row=max_row)
categories = Reference(sheet, min_col=min_column, max_col=min_column, min_row=min_row + 1, max_row=max_row)

bar_chart.add_data(data, titles_from_data=True)
bar_chart.set_categories(categories)

sheet.add_chart(bar_chart, "B12")

bar_chart.title = "Sales by Product line"
bar_chart.style = 5

for i in range(min_column+1, max_column+1):
    letter = get_column_letter(i)
    sheet[f'{letter}{max_row+1}'] = f'=SUM({letter}{min_row+1}:{letter}{max_row})'
    sheet[f'{letter}{max_row+1}'].style = 'Currency'

sheet['A1'] = 'Sales report'
sheet['A2'] = month
sheet['A1'].font = Font('Arial', bold=True, size=20)
sheet['A2'].font = Font('Arial', bold=True, size=10)

wb.save(f'report_{month}.xlsx')